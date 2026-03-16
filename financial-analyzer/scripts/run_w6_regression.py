#!/usr/bin/env python3
"""
W6 最小可用回归检查入口。
"""

import argparse
import datetime
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
TEST_RUNS_DIR = ROOT_DIR / "financial-analyzer" / "test_runs"
ANALYZER_SCRIPT = ROOT_DIR / "financial-analyzer" / "scripts" / "financial_analyzer.py"

CORE_SHEETS = [
    "00_overview",
    "01_kpi_dashboard",
    "02_financial_summary",
    "03_debt_profile",
    "04_liquidity_and_covenants",
    "99_evidence_index",
]

REQUIRED_PAYLOAD_KEYS = [
    "overview",
    "kpi_dashboard",
    "financial_summary",
    "debt_profile",
    "liquidity_and_covenants",
    "evidence_index",
]

CASES = {
    "henglong": {
        "label": "恒隆地产",
        "md_path": ROOT_DIR / "output" / "恒隆地产" / "恒隆地产2024年報" / "恒隆地产2024年報.md",
        "notes_workfile": ROOT_DIR / "financial-analyzer" / "testdata" / "henglong_notes_workfile.json",
        "run_dir": TEST_RUNS_DIR / "w6_henglong",
    },
    "country_garden": {
        "label": "碧桂园",
        "md_path": ROOT_DIR / "cases" / "碧桂园2024年年报分析.md",
        "notes_workfile": ROOT_DIR / "financial-analyzer" / "testdata" / "country_garden_notes_workfile.json",
        "run_dir": TEST_RUNS_DIR / "w6_country_garden",
    },
    "hanghai": {
        "label": "杭海新城控股",
        "md_path": ROOT_DIR / "cases" / "杭海新城控股2024年年报分析.md",
        "notes_workfile": ROOT_DIR / "financial-analyzer" / "testdata" / "hanghai_notes_workfile.json",
        "run_dir": TEST_RUNS_DIR / "w6_hanghai",
    },
}


def now_iso():
    return datetime.datetime.now().astimezone().isoformat(timespec="seconds")


def read_json(path):
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path, payload):
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def read_text(path):
    with open(path, "r", encoding="utf-8") as handle:
        return handle.read()


def tail_text(text, line_limit=40):
    lines = text.strip().splitlines()
    if len(lines) <= line_limit:
        return "\n".join(lines)
    return "\n".join(lines[-line_limit:])


def make_check(name, passed, details=None, errors=None):
    return {
        "name": name,
        "passed": passed,
        "details": details or {},
        "errors": errors or [],
    }


def normalize_path(value):
    return str(Path(value).resolve())


def validate_manifest(run_dir):
    manifest_path = run_dir / "run_manifest.json"
    errors = []
    details = {
        "path": str(manifest_path),
        "exists": manifest_path.exists(),
    }
    manifest = None

    if not manifest_path.exists():
        errors.append("run_manifest.json 未生成")
        return make_check("run_manifest", False, details, errors), manifest

    manifest = read_json(manifest_path)
    details["status"] = manifest.get("status")
    if manifest.get("status") != "success":
        errors.append(f"status 不是 success: {manifest.get('status')!r}")

    artifacts = manifest.get("artifacts") or {}
    required_artifacts = {
        "run_manifest": run_dir / "run_manifest.json",
        "analysis_report": run_dir / "analysis_report.md",
        "soul_export_payload": run_dir / "soul_export_payload.json",
        "financial_output": run_dir / "financial_output.xlsx",
    }
    artifact_details = {}

    for key, expected_path in required_artifacts.items():
        actual_value = artifacts.get(key)
        entry = {
            "expected": str(expected_path.resolve()),
            "actual": actual_value,
            "exists": False,
            "matches_run_dir": False,
        }
        if not actual_value:
            errors.append(f"artifacts.{key} 缺失")
            artifact_details[key] = entry
            continue

        actual_path = Path(actual_value)
        entry["exists"] = actual_path.exists()
        entry["matches_run_dir"] = normalize_path(actual_value) == str(expected_path.resolve())
        artifact_details[key] = entry

        if not entry["exists"]:
            errors.append(f"artifacts.{key} 指向文件不存在: {actual_value}")
        if not entry["matches_run_dir"]:
            errors.append(f"artifacts.{key} 未指向本次 run dir: {actual_value}")

    details["artifacts"] = artifact_details
    return make_check("run_manifest", not errors, details, errors), manifest


def validate_analysis_report(run_dir):
    report_path = run_dir / "analysis_report.md"
    errors = []
    details = {
        "path": str(report_path),
        "exists": report_path.exists(),
    }

    if not report_path.exists():
        errors.append("analysis_report.md 未生成")
        return make_check("analysis_report", False, details, errors)

    content = read_text(report_path)
    details["size_bytes"] = report_path.stat().st_size
    details["contains_run_overview"] = "运行概览" in content
    details["contains_dynamic_focus"] = "动态重点" in content

    if not content.strip():
        errors.append("analysis_report.md 为空")
    if not details["contains_run_overview"]:
        errors.append("analysis_report.md 缺少“运行概览”")
    if not details["contains_dynamic_focus"]:
        errors.append("analysis_report.md 缺少“动态重点”")

    return make_check("analysis_report", not errors, details, errors)


def validate_payload(run_dir):
    payload_path = run_dir / "soul_export_payload.json"
    errors = []
    details = {
        "path": str(payload_path),
        "exists": payload_path.exists(),
    }
    payload = None

    if not payload_path.exists():
        errors.append("soul_export_payload.json 未生成")
        return make_check("soul_export_payload", False, details, errors), payload

    payload = read_json(payload_path)
    details["contract_version"] = payload.get("contract_version")
    details["template_version"] = payload.get("template_version")
    details["has_required_keys"] = {key: key in payload for key in REQUIRED_PAYLOAD_KEYS}
    details["module_manifest_count"] = len(payload.get("module_manifest") or [])
    details["evidence_index_count"] = len(payload.get("evidence_index") or [])

    if payload.get("contract_version") != "soul_export_v1":
        errors.append(f"contract_version 非预期: {payload.get('contract_version')!r}")
    if payload.get("template_version") != "soul_v1_1_alpha":
        errors.append(f"template_version 非预期: {payload.get('template_version')!r}")

    for key in REQUIRED_PAYLOAD_KEYS:
        if key not in payload:
            errors.append(f"payload 缺少固定骨架键: {key}")

    if not payload.get("module_manifest"):
        errors.append("module_manifest 为空")

    return make_check("soul_export_payload", not errors, details, errors), payload


def validate_workbook(run_dir, payload):
    workbook_path = run_dir / "financial_output.xlsx"
    errors = []
    details = {
        "path": str(workbook_path),
        "exists": workbook_path.exists(),
        "sheetnames": [],
        "missing_core_sheets": [],
        "missing_enabled_optional_sheets": [],
    }

    if not workbook_path.exists():
        errors.append("financial_output.xlsx 未生成")
        return make_check("financial_output", False, details, errors)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        details["sheetnames"] = list(workbook.sheetnames)
    finally:
        workbook.close()

    details["missing_core_sheets"] = [
        sheet_name for sheet_name in CORE_SHEETS if sheet_name not in details["sheetnames"]
    ]
    enabled_optional_sheets = []
    if payload:
        for item in payload.get("module_manifest") or []:
            sheet_name = item.get("sheet_name")
            if item.get("enabled") and sheet_name and sheet_name not in CORE_SHEETS:
                enabled_optional_sheets.append(sheet_name)

    details["enabled_optional_sheets"] = enabled_optional_sheets
    details["missing_enabled_optional_sheets"] = [
        sheet_name for sheet_name in enabled_optional_sheets if sheet_name not in details["sheetnames"]
    ]

    if details["missing_core_sheets"]:
        errors.append(f"缺少核心 Sheet: {', '.join(details['missing_core_sheets'])}")
    if details["missing_enabled_optional_sheets"]:
        errors.append(
            "缺少已启用可选 Sheet: "
            + ", ".join(details["missing_enabled_optional_sheets"])
        )

    return make_check("financial_output", not errors, details, errors)


def run_case(case_id, case_config):
    run_dir = case_config["run_dir"]
    if run_dir.exists():
        shutil.rmtree(run_dir)
    run_dir.mkdir(parents=True, exist_ok=True)

    command = [
        sys.executable,
        str(ANALYZER_SCRIPT),
        "--md",
        str(case_config["md_path"]),
        "--notes-workfile",
        str(case_config["notes_workfile"]),
        "--run-dir",
        str(run_dir),
    ]

    started_at = now_iso()
    completed = subprocess.run(
        command,
        cwd=str(ROOT_DIR),
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    finished_at = now_iso()

    manifest_check, manifest = validate_manifest(run_dir)
    report_check = validate_analysis_report(run_dir)
    payload_check, payload = validate_payload(run_dir)
    workbook_check = validate_workbook(run_dir, payload)
    checks = [manifest_check, report_check, payload_check, workbook_check]
    passed = completed.returncode == 0 and all(item["passed"] for item in checks)

    return {
        "case_id": case_id,
        "label": case_config["label"],
        "md_path": str(case_config["md_path"]),
        "notes_workfile": str(case_config["notes_workfile"]),
        "run_dir": str(run_dir),
        "started_at": started_at,
        "finished_at": finished_at,
        "returncode": completed.returncode,
        "passed": passed,
        "command": command,
        "stdout_tail": tail_text(completed.stdout),
        "stderr_tail": tail_text(completed.stderr),
        "checks": checks,
    }


def load_sheetnames(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def collect_known_gaps():
    gaps = []

    legacy_manifest_cases = [
        ("henglong_notes_only", TEST_RUNS_DIR / "henglong_notes_only" / "run_manifest.json"),
        ("henglong_v3", TEST_RUNS_DIR / "henglong_v3" / "run_manifest.json"),
    ]
    windows_path_examples = []
    for case_id, manifest_path in legacy_manifest_cases:
        if not manifest_path.exists():
            continue
        manifest = read_json(manifest_path)
        artifact_path = ((manifest.get("artifacts") or {}).get("run_manifest")) or ""
        if artifact_path.startswith("C:\\"):
            windows_path_examples.append({
                "case_id": case_id,
                "path": artifact_path,
            })

    if windows_path_examples:
        gaps.append({
            "id": "legacy_windows_artifact_paths",
            "category": "observed_repo_state",
            "title": "历史 run_manifest 仍保留 Windows 绝对路径",
            "detail": "历史 test_runs 样本不能直接作为当前 W6 基线，因为 artifacts 仍指向旧环境路径。",
            "evidence": windows_path_examples,
        })

    henglong_v3_manifest = TEST_RUNS_DIR / "henglong_v3" / "run_manifest.json"
    if henglong_v3_manifest.exists():
        manifest = read_json(henglong_v3_manifest)
        if manifest.get("status") is None:
            gaps.append({
                "id": "legacy_manifest_missing_status",
                "category": "observed_repo_state",
                "title": "henglong_v3 缺少正式 status 字段",
                "detail": "该目录可作为历史排障样本，但不能继续充当成功态回归基线。",
                "evidence": [
                    {
                        "case_id": "henglong_v3",
                        "manifest_path": str(henglong_v3_manifest),
                    }
                ],
            })

    legacy_workbook = TEST_RUNS_DIR / "henglong_soul_contract" / "financial_output.xlsx"
    soul_workbook = TEST_RUNS_DIR / "henglong_soul_v1_1_alpha" / "soul_output.xlsx"
    if legacy_workbook.exists() and soul_workbook.exists():
        legacy_sheetnames = load_sheetnames(legacy_workbook)
        soul_sheetnames = load_sheetnames(soul_workbook)
        if "summary" in legacy_sheetnames and "00_overview" in soul_sheetnames:
            gaps.append({
                "id": "mixed_legacy_and_soul_workbooks",
                "category": "observed_repo_state",
                "title": "历史目录同时存在旧内部 workbook 与新 Soul workbook",
                "detail": "W6 v1 只认当前脚本重跑后的 financial_output.xlsx，不直接以历史目录判定通过。",
                "evidence": [
                    {
                        "path": str(legacy_workbook),
                        "sheetnames": legacy_sheetnames,
                    },
                    {
                        "path": str(soul_workbook),
                        "sheetnames": soul_sheetnames,
                    },
                ],
            })

    failure_fixture_manifest = TEST_RUNS_DIR / "missing_notes_workfile" / "run_manifest.json"
    if failure_fixture_manifest.exists():
        manifest = read_json(failure_fixture_manifest)
        if manifest.get("status") == "failed":
            gaps.append({
                "id": "failure_fixture_not_in_happy_path_baseline",
                "category": "known_scope_boundary",
                "title": "失败路径样本已存在，但未纳入本轮 happy-path 基线",
                "detail": "missing_notes_workfile 可作为 W6.1 的失败态回归输入，本轮最小回归仅覆盖 3 个成功案例。",
                "evidence": [
                    {
                        "path": str(failure_fixture_manifest),
                        "failure_reason": manifest.get("failure_reason"),
                    }
                ],
            })

    gaps.extend([
        {
            "id": "visual_review_not_covered",
            "category": "known_scope_boundary",
            "title": "当前最小回归不覆盖 workbook 视觉预览质量",
            "detail": "本轮只验证文件生成和结构约束，不判断预览 PDF/PNG、版式和视觉一致性。",
            "evidence": [],
        },
        {
            "id": "golden_diff_not_covered",
            "category": "known_scope_boundary",
            "title": "当前最小回归不覆盖 JSON/单元格内容级 golden diff",
            "detail": "W6 v1 定义为 smoke regression，不冻结 payload 内容或 Excel 单元格值。",
            "evidence": [],
        },
    ])

    return gaps


def build_summary(results):
    passed_cases = sum(1 for item in results if item["passed"])
    return {
        "selected_case_count": len(results),
        "passed_case_count": passed_cases,
        "failed_case_count": len(results) - passed_cases,
        "all_passed": passed_cases == len(results),
    }


def render_report(results_payload):
    lines = [
        "# W6 Regression Report",
        "",
        f"- 生成时间：{results_payload['generated_at']}",
        f"- 结果 JSON：`{results_payload['results_path']}`",
        f"- 汇总：{results_payload['summary']['passed_case_count']}/{results_payload['summary']['selected_case_count']} 通过",
        "",
        "## 验证范围",
        "",
        "- 固定 3 个案例重跑 `financial_analyzer.py`。",
        "- 仅检查 `run_manifest.json`、`analysis_report.md`、`soul_export_payload.json`、`financial_output.xlsx` 的生成结果和结构约束。",
        "- 不把历史 `test_runs` 目录作为通过依据。",
        "",
        "## 案例结果",
        "",
    ]

    for result in results_payload["results"]:
        status_text = "PASS" if result["passed"] else "FAIL"
        lines.extend([
            f"### {result['label']} (`{result['case_id']}`) - {status_text}",
            "",
            f"- run dir: `{result['run_dir']}`",
            f"- markdown: `{result['md_path']}`",
            f"- notes_workfile: `{result['notes_workfile']}`",
            f"- returncode: `{result['returncode']}`",
            "",
        ])

        for check in result["checks"]:
            marker = "PASS" if check["passed"] else "FAIL"
            lines.append(f"- `{check['name']}`: {marker}")
            for error in check["errors"]:
                lines.append(f"  - {error}")

        if result["stdout_tail"]:
            lines.extend([
                "",
                "```text",
                result["stdout_tail"],
                "```",
            ])

        if result["stderr_tail"]:
            lines.extend([
                "",
                "```text",
                result["stderr_tail"],
                "```",
            ])

        lines.append("")

    lines.extend([
        "## 已知缺口",
        "",
    ])

    for gap in results_payload["known_gaps"]:
        lines.append(f"### {gap['title']}")
        lines.append("")
        lines.append(f"- 分类：`{gap['category']}`")
        lines.append(f"- 说明：{gap['detail']}")
        if gap["evidence"]:
            lines.append("- 证据：")
            for entry in gap["evidence"]:
                lines.append(f"  - `{json.dumps(entry, ensure_ascii=False)}`")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def parse_args():
    parser = argparse.ArgumentParser(description="运行 W6 最小可用回归检查")
    parser.add_argument(
        "--case",
        action="append",
        choices=sorted(CASES.keys()),
        help="仅运行指定案例；可重复传入多个 --case",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    selected_case_ids = args.case or list(CASES.keys())
    results = [run_case(case_id, CASES[case_id]) for case_id in selected_case_ids]
    summary = build_summary(results)
    known_gaps = collect_known_gaps()

    results_path = TEST_RUNS_DIR / "w6_regression_results.json"
    report_path = TEST_RUNS_DIR / "w6_regression_report.md"
    payload = {
        "generated_at": now_iso(),
        "script": str(Path(__file__).resolve()),
        "selected_cases": selected_case_ids,
        "summary": summary,
        "results": results,
        "known_gaps": known_gaps,
        "results_path": str(results_path),
        "report_path": str(report_path),
    }

    write_json(results_path, payload)
    report = render_report(payload)
    with open(report_path, "w", encoding="utf-8") as handle:
        handle.write(report)

    print(f"[INFO] 已写出结果 JSON: {results_path}")
    print(f"[INFO] 已写出结果报告: {report_path}")
    print(
        "[OK] 回归结果: "
        f"{summary['passed_case_count']}/{summary['selected_case_count']} 通过"
    )

    raise SystemExit(0 if summary["all_passed"] else 1)


if __name__ == "__main__":
    main()
